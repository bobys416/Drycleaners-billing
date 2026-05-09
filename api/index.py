import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)) + '/..')

from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import sqlite3, json, smtplib, io, socket
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# For Vercel, use tmp directory for database
DB_PATH = '/tmp/dryclean.db'

# ──────────────────── DATABASE ────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            address TEXT,
            city TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            base_rate REAL NOT NULL,
            description TEXT
        );
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_number TEXT UNIQUE,
            customer_id INTEGER NOT NULL,
            bill_date TEXT DEFAULT (datetime('now','localtime')),
            due_date TEXT,
            subtotal REAL DEFAULT 0,
            discount_pct REAL DEFAULT 0,
            discount_amt REAL DEFAULT 0,
            tax_pct REAL DEFAULT 0,
            tax_amt REAL DEFAULT 0,
            total REAL DEFAULT 0,
            paid_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'Pending',
            notes TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE TABLE IF NOT EXISTS bill_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL,
            cloth_type TEXT,
            service_type TEXT,
            description TEXT,
            quantity INTEGER DEFAULT 1,
            rate REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            FOREIGN KEY (bill_id) REFERENCES bills(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    ''')
    # Seed default services
    c.execute("SELECT COUNT(*) FROM services")
    if c.fetchone()[0] == 0:
        services = [
            ('Dry Cleaning','Formal Wear',150,'Standard dry cleaning'),
            ('Dry Cleaning','Casual Wear',80,'Everyday casuals'),
            ('Dry Cleaning','Ethnic Wear',200,'Sarees, Lehenga, Sherwani'),
            ('Dry Cleaning','Blazer/Jacket',250,'Blazers and jackets'),
            ('Dry Cleaning','Coat/Overcoat',350,'Winter coats'),
            ('Wash & Iron','Shirt',30,'Wash and press shirt'),
            ('Wash & Iron','Trouser',40,'Wash and press trouser'),
            ('Wash & Iron','Saree',80,'Wash and press saree'),
            ('Iron Only','Shirt',15,'Press only'),
            ('Iron Only','Trouser',20,'Press only'),
            ('Stain Removal','Any Cloth',100,'Special stain treatment'),
            ('Alteration','Minor',50,'Minor stitching'),
            ('Alteration','Major',150,'Major alterations'),
        ]
        c.executemany("INSERT INTO services(category,name,base_rate,description) VALUES(?,?,?,?)", services)
    # Default settings
    defaults = [
        ('store_name','FreshClean Dry Cleaners'),
        ('store_address','123 Market Road, Lucknow, UP'),
        ('store_phone','+91 9876543210'),
        ('store_email','store@freshclean.com'),
        ('tax_pct','0'),
        ('smtp_host','smtp.gmail.com'),
        ('smtp_port','587'),
        ('smtp_user',''),
        ('smtp_pass',''),
        ('report_email',''),
    ]
    for k,v in defaults:
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",(k,v))
    conn.commit()
    conn.close()

def next_bill_number():
    conn = get_db()
    row = conn.execute("SELECT bill_number FROM bills ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    if row:
        last = int(row['bill_number'].replace('DC',''))
        return f"DC{last+1:05d}"
    return "DC00001"

# ──────────────────── SETTINGS ────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    conn.close()
    return jsonify({r['key']:r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
def save_settings():
    data = request.json
    conn = get_db()
    for k,v in data.items():
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",(k,str(v)))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── CUSTOMERS ────────────────────
@app.route('/api/customers', methods=['GET'])
def list_customers():
    q = request.args.get('q','')
    conn = get_db()
    if q:
        rows = conn.execute(
            "SELECT * FROM customers WHERE name LIKE ? OR phone LIKE ? OR email LIKE ? ORDER BY name",
            (f'%{q}%',f'%{q}%',f'%{q}%')).fetchall()
    else:
        rows = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/customers', methods=['POST'])
def add_customer():
    d = request.json
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO customers(name,phone,email,address,city,notes) VALUES(?,?,?,?,?,?)",
        (d['name'],d.get('phone',''),d.get('email',''),d.get('address',''),d.get('city',''),d.get('notes','')))
    conn.commit()
    cid = cur.lastrowid
    row = conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/customers/<int:cid>', methods=['GET'])
def get_customer(cid):
    conn = get_db()
    row = conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone()
    conn.close()
    return jsonify(dict(row)) if row else ('Not found',404)

@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    d = request.json
    conn = get_db()
    conn.execute("UPDATE customers SET name=?,phone=?,email=?,address=?,city=?,notes=? WHERE id=?",
        (d['name'],d.get('phone',''),d.get('email',''),d.get('address',''),d.get('city',''),d.get('notes',''),cid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def delete_customer(cid):
    conn = get_db()
    conn.execute("DELETE FROM customers WHERE id=?",(cid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/customers/<int:cid>/bills', methods=['GET'])
def customer_bills(cid):
    conn = get_db()
    rows = conn.execute(
        "SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.customer_id=? ORDER BY b.bill_date DESC",(cid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ──────────────────── SERVICES ────────────────────
@app.route('/api/services', methods=['GET'])
def get_services():
    conn = get_db()
    rows = conn.execute("SELECT * FROM services ORDER BY category,name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/services', methods=['POST'])
def add_service():
    d = request.json
    conn = get_db()
    cur = conn.execute("INSERT INTO services(name,category,base_rate,description) VALUES(?,?,?,?)",
        (d['name'],d.get('category',''),float(d['base_rate']),d.get('description','')))
    conn.commit()
    sid = cur.lastrowid
    row = conn.execute("SELECT * FROM services WHERE id=?",(sid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/services/<int:sid>', methods=['PUT'])
def update_service(sid):
    d = request.json
    conn = get_db()
    conn.execute("UPDATE services SET name=?,category=?,base_rate=?,description=? WHERE id=?",
        (d['name'],d.get('category',''),float(d['base_rate']),d.get('description',''),sid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/services/<int:sid>', methods=['DELETE'])
def delete_service(sid):
    conn = get_db()
    conn.execute("DELETE FROM services WHERE id=?",(sid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── BILLS ────────────────────
@app.route('/api/bills', methods=['GET'])
def list_bills():
    conn = get_db()
    rows = conn.execute(
        "SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id ORDER BY b.bill_date DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bills', methods=['POST'])
def create_bill():
    d = request.json
    conn = get_db()
    bill_no = next_bill_number()
    cur = conn.execute(
        "INSERT INTO bills(bill_number,customer_id,bill_date,due_date,subtotal,discount_pct,discount_amt,tax_pct,tax_amt,total,status) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
        (bill_no, d['customer_id'], d.get('bill_date',datetime.now().strftime('%Y-%m-%d')), d.get('due_date',''),
         d.get('subtotal',0), d.get('discount_pct',0), d.get('discount_amt',0), d.get('tax_pct',0),
         d.get('tax_amt',0), d.get('total',0), 'Pending'))
    conn.commit()
    bid = cur.lastrowid
    row = conn.execute("SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.id=?",(bid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/bills/<int:bid>', methods=['GET'])
def get_bill(bid):
    conn = get_db()
    bill = conn.execute("SELECT b.*, c.name as customer_name, c.phone, c.email FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.id=?",(bid,)).fetchone()
    items = conn.execute("SELECT * FROM bill_items WHERE bill_id=?",(bid,)).fetchall()
    conn.close()
    if not bill: return ('Not found',404)
    return jsonify({**dict(bill), 'items': [dict(i) for i in items]})

@app.route('/api/bills/<int:bid>', methods=['PUT'])
def update_bill(bid):
    d = request.json
    conn = get_db()
    conn.execute(
        "UPDATE bills SET subtotal=?,discount_pct=?,discount_amt=?,tax_pct=?,tax_amt=?,total=?,status=?,notes=? WHERE id=?",
        (d['subtotal'],d['discount_pct'],d['discount_amt'],d['tax_pct'],d['tax_amt'],d['total'],d.get('status','Pending'),d.get('notes',''),bid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/bills/<int:bid>/pay', methods=['POST'])
def pay_bill(bid):
    d = request.json
    conn = get_db()
    amount = float(d['amount'])
    bill = conn.execute("SELECT * FROM bills WHERE id=?",(bid,)).fetchone()
    new_paid = bill['paid_amount'] + amount
    new_status = 'Paid' if new_paid >= bill['total'] else 'Partial'
    conn.execute("UPDATE bills SET paid_amount=?,status=? WHERE id=?",(new_paid,new_status,bid))
    conn.commit(); conn.close()
    return jsonify({'ok':True,'paid_amount':new_paid})

@app.route('/api/bills/<int:bid>', methods=['DELETE'])
def delete_bill(bid):
    conn = get_db()
    conn.execute("DELETE FROM bills WHERE id=?",(bid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── BILL ITEMS ────────────────────
@app.route('/api/bills/<int:bid>/items', methods=['POST'])
def add_bill_item(bid):
    d = request.json
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO bill_items(bill_id,cloth_type,service_type,description,quantity,rate,amount) VALUES(?,?,?,?,?,?,?)",
        (bid, d['cloth_type'], d['service_type'], d.get('description',''), d['quantity'], d['rate'], d['amount']))
    conn.commit()
    iid = cur.lastrowid
    row = conn.execute("SELECT * FROM bill_items WHERE id=?",(iid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/bill-items/<int:iid>', methods=['DELETE'])
def delete_bill_item(iid):
    conn = get_db()
    conn.execute("DELETE FROM bill_items WHERE id=?",(iid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── DASHBOARD ────────────────────
@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    month = request.args.get('month', datetime.now().strftime('%Y-%m'))
    conn = get_db()
    
    # Total revenue
    revenue = conn.execute(
        "SELECT SUM(total) as total FROM bills WHERE bill_date LIKE ?",(f'{month}%',)).fetchone()['total'] or 0
    
    # Pending dues
    dues = conn.execute(
        "SELECT SUM(total - paid_amount) as dues FROM bills WHERE status != 'Paid'"
    ).fetchone()['dues'] or 0
    
    # Recent bills
    recent = conn.execute(
        "SELECT b.*, c.name FROM bills b JOIN customers c ON b.customer_id=c.id ORDER BY b.bill_date DESC LIMIT 5"
    ).fetchall()
    
    conn.close()
    return jsonify({
        'month': month,
        'revenue': revenue,
        'pending_dues': dues,
        'recent_bills': [dict(r) for r in recent]
    })

# ──────────────────── EXCEL EXPORT ────────────────────
def build_excel(month=None, customer_id=None):
    conn = get_db()
    query = "SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id"
    params = []
    if month:
        query += " WHERE b.bill_date LIKE ?"
        params.append(f'{month}%')
    if customer_id:
        query += (" AND " if month else " WHERE ") + "b.customer_id=?"
        params.append(customer_id)
    bills = conn.execute(query, params).fetchall()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    store_name = conn.execute("SELECT value FROM settings WHERE key='store_name'").fetchone()['value']
    
    # Styling
    title_font = Font(name='Calibri',size=14,bold=True,color='1A3C5E')
    header_font = Font(name='Calibri',size=11,bold=True,color='ffffff')
    header_fill = PatternFill(start_color='0F2744',end_color='0F2744',fill_type='solid')
    alt_fill = PatternFill(start_color='E8F0F8',end_color='E8F0F8',fill_type='solid')
    money_fill = PatternFill(start_color='D4E8F7',end_color='D4E8F7',fill_type='solid')
    border = Border(left=Side(style='thin',color='D6E2F0'),right=Side(style='thin',color='D6E2F0'),
                    top=Side(style='thin',color='D6E2F0'),bottom=Side(style='thin',color='D6E2F0'))

    # ── Sheet 1: Bills Summary ──
    ws1 = wb.create_sheet("Bills Summary")
    ws1.merge_cells('A1:L1')
    ws1['A1'] = f"{store_name} — Bills Summary {month or 'Report'}"
    ws1['A1'].font = title_font; ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 28
    h1 = ['Bill No','Date','Customer','Phone','Amount','Discount','Tax','Total','Paid','Status','Notes','']
    for col, h in enumerate(h1, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    w1 = [12,12,22,15,12,10,10,12,12,12,20,1]
    for i,w in enumerate(w1,1): ws1.column_dimensions[get_column_letter(i)].width = w

    for idx, b in enumerate(bills, 3):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        vals = [b['bill_number'], b['bill_date'][:10] if b['bill_date'] else '', b['customer_name'],
                b['phone'] or '', b['total'], b['discount_amt'], b['tax_amt'], b['total'],
                b['paid_amount'], b['status'], b['notes'] or '', '']
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=idx, column=col, value=val)
            cell.fill = fill; cell.border = border
            if col in [5,6,7,8,9]: cell.number_format = '₹#,##0.00'

    # Totals row
    total_row = len(bills) + 3
    ws1.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=total_row, column=12, value=sum(b['total'] for b in bills)).number_format = '₹#,##0.00'
    ws1.cell(row=total_row, column=12).font = Font(bold=True, color="1A3C5E")
    ws1.cell(row=total_row, column=12).fill = money_fill

    # ── Sheet 2: Itemized Details ──
    ws2 = wb.create_sheet("Itemized Details")
    ws2.merge_cells('A1:J1')
    ws2['A1'] = f"{store_name} — Itemized Bill Details"
    ws2['A1'].font = title_font; ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 28
    h2 = ['Bill No','Date','Customer','Cloth Type','Service Type','Description','Qty','Rate','Amount','Status']
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    w2 = [12,12,22,18,20,25,6,10,12,12]
    for i,w in enumerate(w2,1): ws2.column_dimensions[get_column_letter(i)].width = w

    r = 3
    for b in bills:
        items = conn.execute("SELECT * FROM bill_items WHERE bill_id=?",(b['id'],)).fetchall()
        for idx, item in enumerate(items):
            fill = alt_fill if r % 2 == 0 else PatternFill()
            vals = [b['bill_number'] if idx==0 else '', b['bill_date'][:10] if (b['bill_date'] and idx==0) else '',
                    b['customer_name'] if idx==0 else '', item['cloth_type'], item['service_type'],
                    item['description'], item['quantity'], item['rate'], item['amount'],
                    b['status'] if idx==0 else '']
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.fill = fill; cell.border = border
                if col in [8,9]: cell.number_format = '₹#,##0.00'
            r += 1

    # ── Sheet 3: Customer Ledger ──
    ws3 = wb.create_sheet("Customer Ledger")
    ws3.merge_cells('A1:G1')
    ws3['A1'] = f"{store_name} — Customer Ledger"
    ws3['A1'].font = title_font; ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 28
    h3 = ['Customer','Phone','Email','City','Total Bills','Total Amount','Outstanding']
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    w3 = [25,15,28,15,12,15,15]
    for i,w in enumerate(w3,1): ws3.column_dimensions[get_column_letter(i)].width = w

    cust_data = conn.execute("""
        SELECT c.name, c.phone, c.email, c.city,
               COUNT(b.id) as bill_count,
               COALESCE(SUM(b.total),0) as total_amt,
               COALESCE(SUM(b.total - b.paid_amount),0) as outstanding
        FROM customers c LEFT JOIN bills b ON c.id=b.customer_id
        GROUP BY c.id ORDER BY total_amt DESC
    """).fetchall()
    for idx, row in enumerate(cust_data, 3):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        vals = [row['name'], row['phone'], row['email'], row['city'], row['bill_count'], row['total_amt'], row['outstanding']]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=idx, column=col, value=val)
            cell.fill = fill; cell.border = border
            if col in [6,7]: cell.number_format = '₹#,##0.00'

    conn.close()
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    month = request.args.get('month')
    cid   = request.args.get('customer_id')
    buf   = build_excel(month=month, customer_id=cid)
    fname = f"dryclean_{month or 'all'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ──────────────────── EMAIL REPORT ────────────────────
@app.route('/api/send-report', methods=['POST'])
def send_report():
    d = request.json
    month = d.get('month', datetime.now().strftime('%Y-%m'))
    to_email = d.get('email')
    conn = get_db()
    settings = {r['key']:r['value'] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    conn.close()
    if not to_email:
        to_email = settings.get('report_email','')
    if not to_email:
        return jsonify({'ok':False,'error':'No recipient email set'}), 400
    smtp_host = settings.get('smtp_host','smtp.gmail.com')
    smtp_port = int(settings.get('smtp_port','587'))
    smtp_user = settings.get('smtp_user','')
    smtp_pass = settings.get('smtp_pass','')
    if not smtp_user:
        return jsonify({'ok':False,'error':'SMTP not configured in Settings'}), 400
    try:
        buf = build_excel(month=month)
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To']   = to_email
        msg['Subject'] = f"{settings.get('store_name','DryClean')} — Monthly Report {month}"
        body = f"""Dear Team,

Please find attached the monthly billing report for {month}.

Store: {settings.get('store_name','')}
Report Period: {month}

This report includes:
- Bills Summary
- Itemized Details  
- Customer Ledger

Regards,
{settings.get('store_name','')} Billing System
"""
        msg.attach(MIMEText(body,'plain'))
        part = MIMEBase('application','octet-stream')
        part.set_payload(buf.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="report_{month}.xlsx"')
        msg.attach(part)
        
        # Set socket timeout for Vercel environment
        socket.setdefaulttimeout(10)
        
        with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return jsonify({'ok':True})
    except socket.gaierror:
        return jsonify({'ok':False,'error':f'DNS error: Cannot resolve hostname {smtp_host}. Check your SMTP settings.'}), 500
    except socket.timeout:
        return jsonify({'ok':False,'error':'Connection timeout. SMTP server took too long to respond. Check your internet connection.'}), 500
    except smtplib.SMTPAuthenticationError:
        return jsonify({'ok':False,'error':'SMTP authentication failed. Check your email and password in Settings.'}), 400
    except Exception as e:
        error_msg = str(e)
        if 'Name or service not known' in error_msg or 'nodename nor servname provided' in error_msg:
            return jsonify({'ok':False,'error':f'DNS resolution failed. Check SMTP host: {smtp_host}'}), 500
        return jsonify({'ok':False,'error':f'Email send failed: {error_msg}'}), 500

# ──────────────────── SERVE FRONTEND ────────────────────
@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint for Vercel"""
    return jsonify({'status': 'ok', 'app': 'FreshClean Billing System'}), 200

@app.route('/')
def index():
    try:
        return render_template('index.html')
    except Exception as e:
        # Fallback if template not found (shouldn't happen but safe)
        return jsonify({'error': 'Frontend not available', 'details': str(e)}), 500

# Initialize database on first request (lazy initialization for Vercel)
_db_initialized = False

@app.before_request
def ensure_db_initialized():
    global _db_initialized
    if not _db_initialized:
        try:
            init_db()
            _db_initialized = True
        except Exception as e:
            print(f"Database initialization warning: {e}")

# Export app for Vercel
export = app
