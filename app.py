from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
import sqlite3, os, json, smtplib, io
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'dryclean.db')

# ──────────────────── DATABASE ────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.executescript('''
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            phone TEXT,
            email TEXT,
            address TEXT,
            city TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS services (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT,
            base_rate REAL NOT NULL,
            description TEXT
        );
        CREATE TABLE IF NOT EXISTS bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_number TEXT UNIQUE,
            customer_id INTEGER NOT NULL,
            bill_date TEXT DEFAULT (datetime('now','localtime')),
            due_date TEXT,
            subtotal REAL DEFAULT 0,
            discount_pct REAL DEFAULT 0,
            discount_amt REAL DEFAULT 0,
            tax_pct REAL DEFAULT 0,
            tax_amt REAL DEFAULT 0,
            total REAL DEFAULT 0,
            paid_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'Pending',
            notes TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );
        CREATE TABLE IF NOT EXISTS bill_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL,
            cloth_type TEXT,
            service_type TEXT,
            description TEXT,
            quantity INTEGER DEFAULT 1,
            rate REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            FOREIGN KEY (bill_id) REFERENCES bills(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        );
    ''')
    # Seed default services
    c.execute("SELECT COUNT(*) FROM services")
    if c.fetchone()[0] == 0:
        services = [
            ('Dry Cleaning','Formal Wear',150,'Standard dry cleaning'),
            ('Dry Cleaning','Casual Wear',80,'Everyday casuals'),
            ('Dry Cleaning','Ethnic Wear',200,'Sarees, Lehenga, Sherwani'),
            ('Dry Cleaning','Blazer/Jacket',250,'Blazers and jackets'),
            ('Dry Cleaning','Coat/Overcoat',350,'Winter coats'),
            ('Wash & Iron','Shirt',30,'Wash and press shirt'),
            ('Wash & Iron','Trouser',40,'Wash and press trouser'),
            ('Wash & Iron','Saree',80,'Wash and press saree'),
            ('Iron Only','Shirt',15,'Press only'),
            ('Iron Only','Trouser',20,'Press only'),
            ('Stain Removal','Any Cloth',100,'Special stain treatment'),
            ('Alteration','Minor',50,'Minor stitching'),
            ('Alteration','Major',150,'Major alterations'),
        ]
        c.executemany("INSERT INTO services(category,name,base_rate,description) VALUES(?,?,?,?)", services)
    # Default settings
    defaults = [
        ('store_name','FreshClean Dry Cleaners'),
        ('store_address','123 Market Road, Lucknow, UP'),
        ('store_phone','+91 9876543210'),
        ('store_email','store@freshclean.com'),
        ('tax_pct','0'),
        ('smtp_host','smtp.gmail.com'),
        ('smtp_port','587'),
        ('smtp_user',''),
        ('smtp_pass',''),
        ('report_email',''),
    ]
    for k,v in defaults:
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)",(k,v))
    conn.commit()
    conn.close()

def next_bill_number():
    conn = get_db()
    row = conn.execute("SELECT bill_number FROM bills ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    if row:
        last = int(row['bill_number'].replace('DC',''))
        return f"DC{last+1:05d}"
    return "DC00001"

# ──────────────────── SETTINGS ────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    conn = get_db()
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    conn.close()
    return jsonify({r['key']:r['value'] for r in rows})

@app.route('/api/settings', methods=['POST'])
def save_settings():
    data = request.json
    conn = get_db()
    for k,v in data.items():
        conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)",(k,str(v)))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── CUSTOMERS ────────────────────
@app.route('/api/customers', methods=['GET'])
def list_customers():
    q = request.args.get('q','')
    conn = get_db()
    if q:
        rows = conn.execute(
            "SELECT * FROM customers WHERE name LIKE ? OR phone LIKE ? OR email LIKE ? ORDER BY name",
            (f'%{q}%',f'%{q}%',f'%{q}%')).fetchall()
    else:
        rows = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/customers', methods=['POST'])
def add_customer():
    d = request.json
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO customers(name,phone,email,address,city,notes) VALUES(?,?,?,?,?,?)",
        (d['name'],d.get('phone',''),d.get('email',''),d.get('address',''),d.get('city',''),d.get('notes','')))
    conn.commit()
    cid = cur.lastrowid
    row = conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/customers/<int:cid>', methods=['GET'])
def get_customer(cid):
    conn = get_db()
    row = conn.execute("SELECT * FROM customers WHERE id=?",(cid,)).fetchone()
    conn.close()
    return jsonify(dict(row)) if row else ('Not found',404)

@app.route('/api/customers/<int:cid>', methods=['PUT'])
def update_customer(cid):
    d = request.json
    conn = get_db()
    conn.execute("UPDATE customers SET name=?,phone=?,email=?,address=?,city=?,notes=? WHERE id=?",
        (d['name'],d.get('phone',''),d.get('email',''),d.get('address',''),d.get('city',''),d.get('notes',''),cid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/customers/<int:cid>', methods=['DELETE'])
def delete_customer(cid):
    conn = get_db()
    conn.execute("DELETE FROM customers WHERE id=?",(cid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/customers/<int:cid>/bills', methods=['GET'])
def customer_bills(cid):
    conn = get_db()
    rows = conn.execute(
        "SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.customer_id=? ORDER BY b.bill_date DESC",(cid,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ──────────────────── SERVICES ────────────────────
@app.route('/api/services', methods=['GET'])
def list_services():
    conn = get_db()
    rows = conn.execute("SELECT * FROM services ORDER BY category,name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/services', methods=['POST'])
def add_service():
    d = request.json
    conn = get_db()
    cur = conn.execute("INSERT INTO services(name,category,base_rate,description) VALUES(?,?,?,?)",
        (d['name'],d.get('category',''),float(d['base_rate']),d.get('description','')))
    conn.commit()
    row = conn.execute("SELECT * FROM services WHERE id=?",(cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/services/<int:sid>', methods=['PUT'])
def update_service(sid):
    d = request.json
    conn = get_db()
    conn.execute("UPDATE services SET name=?,category=?,base_rate=?,description=? WHERE id=?",
        (d['name'],d.get('category',''),float(d['base_rate']),d.get('description',''),sid))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/services/<int:sid>', methods=['DELETE'])
def delete_service(sid):
    conn = get_db()
    conn.execute("DELETE FROM services WHERE id=?",(sid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── BILLS ────────────────────
@app.route('/api/bills', methods=['GET'])
def list_bills():
    month = request.args.get('month')
    cid   = request.args.get('customer_id')
    status= request.args.get('status')
    conn  = get_db()
    q = "SELECT b.*, c.name as customer_name, c.phone as customer_phone FROM bills b JOIN customers c ON b.customer_id=c.id WHERE 1=1"
    params = []
    if month:
        q += " AND strftime('%Y-%m', b.bill_date)=?"; params.append(month)
    if cid:
        q += " AND b.customer_id=?"; params.append(cid)
    if status:
        q += " AND b.status=?"; params.append(status)
    q += " ORDER BY b.bill_date DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/bills', methods=['POST'])
def create_bill():
    d = request.json
    subtotal = sum(float(i.get('amount',0)) for i in d.get('items',[]))
    disc_pct  = float(d.get('discount_pct',0))
    disc_amt  = round(subtotal * disc_pct/100, 2)
    tax_pct   = float(d.get('tax_pct',0))
    tax_amt   = round((subtotal - disc_amt) * tax_pct/100, 2)
    total     = round(subtotal - disc_amt + tax_amt, 2)
    conn = get_db()
    bnum = next_bill_number()
    cur = conn.execute(
        "INSERT INTO bills(bill_number,customer_id,bill_date,due_date,subtotal,discount_pct,discount_amt,tax_pct,tax_amt,total,paid_amount,status,notes) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (bnum, d['customer_id'], d.get('bill_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
         d.get('due_date',''), subtotal, disc_pct, disc_amt, tax_pct, tax_amt, total,
         float(d.get('paid_amount',0)), d.get('status','Pending'), d.get('notes','')))
    bid = cur.lastrowid
    for item in d.get('items',[]):
        conn.execute("INSERT INTO bill_items(bill_id,cloth_type,service_type,description,quantity,rate,amount) VALUES(?,?,?,?,?,?,?)",
            (bid, item.get('cloth_type',''), item.get('service_type',''), item.get('description',''),
             int(item.get('quantity',1)), float(item.get('rate',0)), float(item.get('amount',0))))
    conn.commit()
    row = conn.execute("SELECT b.*, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.id=?",(bid,)).fetchone()
    conn.close()
    return jsonify(dict(row))

@app.route('/api/bills/<int:bid>', methods=['GET'])
def get_bill(bid):
    conn = get_db()
    bill = conn.execute("SELECT b.*, c.name as customer_name, c.phone as customer_phone, c.email as customer_email, c.address as customer_address, c.city as customer_city FROM bills b JOIN customers c ON b.customer_id=c.id WHERE b.id=?",(bid,)).fetchone()
    items = conn.execute("SELECT * FROM bill_items WHERE bill_id=?",(bid,)).fetchall()
    conn.close()
    if not bill: return ('Not found',404)
    return jsonify({'bill':dict(bill), 'items':[dict(i) for i in items]})

@app.route('/api/bills/<int:bid>', methods=['PUT'])
def update_bill(bid):
    d = request.json
    conn = get_db()
    if 'status' in d and len(d)==1:
        conn.execute("UPDATE bills SET status=? WHERE id=?",(d['status'],bid))
    elif 'paid_amount' in d:
        conn.execute("UPDATE bills SET paid_amount=?,status=? WHERE id=?",(d['paid_amount'],d.get('status','Partial'),bid))
    else:
        items = d.get('items',[])
        subtotal = sum(float(i.get('amount',0)) for i in items)
        disc_pct = float(d.get('discount_pct',0))
        disc_amt = round(subtotal * disc_pct/100, 2)
        tax_pct  = float(d.get('tax_pct',0))
        tax_amt  = round((subtotal-disc_amt)*tax_pct/100, 2)
        total    = round(subtotal - disc_amt + tax_amt, 2)
        conn.execute("UPDATE bills SET customer_id=?,bill_date=?,due_date=?,subtotal=?,discount_pct=?,discount_amt=?,tax_pct=?,tax_amt=?,total=?,paid_amount=?,status=?,notes=? WHERE id=?",
            (d['customer_id'],d.get('bill_date'),d.get('due_date',''),subtotal,disc_pct,disc_amt,tax_pct,tax_amt,total,
             float(d.get('paid_amount',0)),d.get('status','Pending'),d.get('notes',''),bid))
        conn.execute("DELETE FROM bill_items WHERE bill_id=?",(bid,))
        for item in items:
            conn.execute("INSERT INTO bill_items(bill_id,cloth_type,service_type,description,quantity,rate,amount) VALUES(?,?,?,?,?,?,?)",
                (bid,item.get('cloth_type',''),item.get('service_type',''),item.get('description',''),
                 int(item.get('quantity',1)),float(item.get('rate',0)),float(item.get('amount',0))))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

@app.route('/api/bills/<int:bid>', methods=['DELETE'])
def delete_bill(bid):
    conn = get_db()
    conn.execute("DELETE FROM bills WHERE id=?",(bid,))
    conn.commit(); conn.close()
    return jsonify({'ok':True})

# ──────────────────── DASHBOARD STATS ────────────────────
@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    conn = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    month = datetime.now().strftime('%Y-%m')
    stats = {}
    stats['total_customers'] = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    stats['total_bills']     = conn.execute("SELECT COUNT(*) FROM bills").fetchone()[0]
    stats['monthly_revenue'] = conn.execute("SELECT COALESCE(SUM(total),0) FROM bills WHERE strftime('%Y-%m',bill_date)=? AND status != 'Cancelled'",(month,)).fetchone()[0]
    stats['pending_amount']  = conn.execute("SELECT COALESCE(SUM(total-paid_amount),0) FROM bills WHERE status IN ('Pending','Partial')").fetchone()[0]
    stats['today_bills']     = conn.execute("SELECT COUNT(*) FROM bills WHERE date(bill_date)=?",(today,)).fetchone()[0]
    stats['today_revenue']   = conn.execute("SELECT COALESCE(SUM(total),0) FROM bills WHERE date(bill_date)=? AND status != 'Cancelled'",(today,)).fetchone()[0]
    recent = conn.execute("SELECT b.bill_number, b.total, b.status, b.bill_date, c.name as customer_name FROM bills b JOIN customers c ON b.customer_id=c.id ORDER BY b.bill_date DESC LIMIT 8").fetchall()
    stats['recent_bills'] = [dict(r) for r in recent]
    monthly = conn.execute("SELECT strftime('%Y-%m',bill_date) as m, COALESCE(SUM(total),0) as rev FROM bills WHERE status != 'Cancelled' GROUP BY m ORDER BY m DESC LIMIT 6").fetchall()
    stats['monthly_chart'] = [dict(r) for r in monthly]
    top_services = conn.execute("SELECT service_type, COUNT(*) as cnt, SUM(amount) as rev FROM bill_items GROUP BY service_type ORDER BY rev DESC LIMIT 5").fetchall()
    stats['top_services'] = [dict(r) for r in top_services]
    conn.close()
    return jsonify(stats)

# ──────────────────── EXCEL EXPORT ────────────────────
def build_excel(month=None, customer_id=None):
    conn = get_db()
    q = "SELECT b.*, c.name as customer_name, c.phone, c.email, c.address, c.city FROM bills b JOIN customers c ON b.customer_id=c.id WHERE 1=1"
    params = []
    if month:
        q += " AND strftime('%Y-%m', b.bill_date)=?"; params.append(month)
    if customer_id:
        q += " AND b.customer_id=?"; params.append(customer_id)
    q += " ORDER BY b.bill_date DESC"
    bills = conn.execute(q, params).fetchall()

    wb = openpyxl.Workbook()
    # ── Sheet 1: Bills Summary ──
    ws1 = wb.active
    ws1.title = "Bills Summary"
    header_fill   = PatternFill("solid", fgColor="1A3C5E")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    alt_fill      = PatternFill("solid", fgColor="EBF2FA")
    money_fill    = PatternFill("solid", fgColor="E8F5E9")
    title_font    = Font(bold=True, size=14, color="1A3C5E")
    thin          = Side(border_style="thin", color="CCCCCC")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    store_name = conn.execute("SELECT value FROM settings WHERE key='store_name'").fetchone()
    store_name = store_name['value'] if store_name else "DryClean Store"
    ws1.merge_cells('A1:M1')
    ws1['A1'] = f"{store_name} — Bill Records" + (f" ({month})" if month else "")
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.row_dimensions[1].height = 28

    headers = ['Bill No','Date','Customer','Phone','Email','Address','City','Subtotal','Discount%','Disc Amt','Tax Amt','Total','Status']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border

    col_widths = [12,18,22,15,25,30,15,12,11,11,10,12,12]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    for row_idx, b in enumerate(bills, 3):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        vals = [b['bill_number'], b['bill_date'][:10] if b['bill_date'] else '',
                b['customer_name'], b['phone'], b['email'], b['address'], b['city'],
                b['subtotal'], b['discount_pct'], b['discount_amt'], b['tax_amt'], b['total'], b['status']]
        for col, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            cell.fill = fill; cell.border = border
            cell.alignment = Alignment(horizontal='center' if col in [1,2,9,13] else 'left')
            if col in [8,10,11,12]:
                cell.number_format = '₹#,##0.00'

    # Totals row
    total_row = len(bills) + 3
    ws1.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=total_row, column=12, value=sum(b['total'] for b in bills)).number_format = '₹#,##0.00'
    ws1.cell(row=total_row, column=12).font = Font(bold=True, color="1A3C5E")
    ws1.cell(row=total_row, column=12).fill = money_fill

    # ── Sheet 2: Itemized Details ──
    ws2 = wb.create_sheet("Itemized Details")
    ws2.merge_cells('A1:J1')
    ws2['A1'] = f"{store_name} — Itemized Bill Details"
    ws2['A1'].font = title_font; ws2['A1'].alignment = Alignment(horizontal='center')
    ws2.row_dimensions[1].height = 28
    h2 = ['Bill No','Date','Customer','Cloth Type','Service Type','Description','Qty','Rate','Amount','Status']
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    w2 = [12,12,22,18,20,25,6,10,12,12]
    for i,w in enumerate(w2,1): ws2.column_dimensions[get_column_letter(i)].width = w

    r = 3
    for b in bills:
        items = conn.execute("SELECT * FROM bill_items WHERE bill_id=?",(b['id'],)).fetchall()
        for idx, item in enumerate(items):
            fill = alt_fill if r % 2 == 0 else PatternFill()
            vals = [b['bill_number'] if idx==0 else '', b['bill_date'][:10] if (b['bill_date'] and idx==0) else '',
                    b['customer_name'] if idx==0 else '', item['cloth_type'], item['service_type'],
                    item['description'], item['quantity'], item['rate'], item['amount'],
                    b['status'] if idx==0 else '']
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=col, value=val)
                cell.fill = fill; cell.border = border
                if col in [8,9]: cell.number_format = '₹#,##0.00'
            r += 1

    # ── Sheet 3: Customer Ledger ──
    ws3 = wb.create_sheet("Customer Ledger")
    ws3.merge_cells('A1:G1')
    ws3['A1'] = f"{store_name} — Customer Ledger"
    ws3['A1'].font = title_font; ws3['A1'].alignment = Alignment(horizontal='center')
    ws3.row_dimensions[1].height = 28
    h3 = ['Customer','Phone','Email','City','Total Bills','Total Amount','Outstanding']
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal='center'); cell.border = border
    w3 = [25,15,28,15,12,15,15]
    for i,w in enumerate(w3,1): ws3.column_dimensions[get_column_letter(i)].width = w

    cust_data = conn.execute("""
        SELECT c.name, c.phone, c.email, c.city,
               COUNT(b.id) as bill_count,
               COALESCE(SUM(b.total),0) as total_amt,
               COALESCE(SUM(b.total - b.paid_amount),0) as outstanding
        FROM customers c LEFT JOIN bills b ON c.id=b.customer_id
        GROUP BY c.id ORDER BY total_amt DESC
    """).fetchall()
    for idx, row in enumerate(cust_data, 3):
        fill = alt_fill if idx % 2 == 0 else PatternFill()
        vals = [row['name'], row['phone'], row['email'], row['city'], row['bill_count'], row['total_amt'], row['outstanding']]
        for col, val in enumerate(vals, 1):
            cell = ws3.cell(row=idx, column=col, value=val)
            cell.fill = fill; cell.border = border
            if col in [6,7]: cell.number_format = '₹#,##0.00'

    conn.close()
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    month = request.args.get('month')
    cid   = request.args.get('customer_id')
    buf   = build_excel(month=month, customer_id=cid)
    fname = f"dryclean_{month or 'all'}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ──────────────────── EMAIL REPORT ────────────────────
@app.route('/api/send-report', methods=['POST'])
def send_report():
    d = request.json
    month = d.get('month', datetime.now().strftime('%Y-%m'))
    to_email = d.get('email')
    conn = get_db()
    settings = {r['key']:r['value'] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    conn.close()
    if not to_email:
        to_email = settings.get('report_email','')
    if not to_email:
        return jsonify({'ok':False,'error':'No recipient email set'}), 400
    smtp_host = settings.get('smtp_host','smtp.gmail.com')
    smtp_port = int(settings.get('smtp_port','587'))
    smtp_user = settings.get('smtp_user','')
    smtp_pass = settings.get('smtp_pass','')
    if not smtp_user:
        return jsonify({'ok':False,'error':'SMTP not configured in Settings'}), 400
    try:
        buf = build_excel(month=month)
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To']   = to_email
        msg['Subject'] = f"{settings.get('store_name','DryClean')} — Monthly Report {month}"
        body = f"""Dear Team,

Please find attached the monthly billing report for {month}.

Store: {settings.get('store_name','')}
Report Period: {month}

This report includes:
- Bills Summary
- Itemized Details  
- Customer Ledger

Regards,
{settings.get('store_name','')} Billing System
"""
        msg.attach(MIMEText(body,'plain'))
        part = MIMEBase('application','octet-stream')
        part.set_payload(buf.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="report_{month}.xlsx"')
        msg.attach(part)
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ──────────────────── SERVE FRONTEND ────────────────────
@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    init_db()
    print("\n✅ FreshClean Billing System running at http://localhost:5055\n")
    app.run(debug=False, port=5055, host='0.0.0.0')
